"""
common.library.export.exporter — DataExporter class (L22).

Universal CSV + Excel exporter for pipeline results. Accepts ``list[dict]``
rows directly — no pandas dependency. Uses ``csv`` stdlib for CSV output
and ``openpyxl`` for Excel output.

Error codes (S-DE-* range):
    S-DE-001 — CSV write failed (I/O error)
    S-DE-002 — Excel write failed (I/O or format error)
    S-DE-003 — Output directory not writable
    S-DE-004 — File exists and overwrite disabled

Revision: 0.1
Date: 2026-07-18
Author: opencode
Summary: Initial implementation — export_to_csv (BOM UTF-8, DictWriter),
         export_to_excel (auto-column-width, bold+frozen headers),
         export_multi_sheet (multi-sheet workbook).
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, List, Optional


# ---------------------------------------------------------------------------
# L10 — Error codes (S-DE-* range: System DataExport)
# ---------------------------------------------------------------------------
_EXPORT_ERRORS = {
    "S-DE-001": "CSV write failed — {detail}",
    "S-DE-002": "Excel write failed — {detail}",
    "S-DE-003": "Output directory not writable: {path}",
    "S-DE-004": "File exists and overwrite disabled: {path}",
}


class DataExportError(Exception):
    """Structured export error with S-DE-* code."""

    def __init__(self, code: str, detail: str = "", path: Optional[Path] = None) -> None:
        template = _EXPORT_ERRORS.get(code, "{detail}")
        msg = template.format(detail=detail, path=str(path) if path else "")
        super().__init__(f"[{code}] {msg}")
        self.code = code
        self.detail = detail
        self.path = path


# ---------------------------------------------------------------------------
# DataExporter
# ---------------------------------------------------------------------------

class DataExporter:
    """Universal CSV + Excel exporter for ``list[dict]`` rows.

    Args:
        overwrite: If False, raise S-DE-004 when output file already exists.
                   Default True (consistent with I124 design — overwrite on each run).
    """

    def __init__(self, overwrite: bool = True) -> None:
        self.overwrite = overwrite

    # ------------------------------------------------------------------
    # CSV export
    # ------------------------------------------------------------------

    def export_to_csv(
        self,
        rows: List[Dict[str, Any]],
        path: Path,
        columns: Optional[List[str]] = None,
    ) -> Path:
        """Write *rows* as a CSV file with BOM for Excel UTF-8 compatibility.

        Args:
            rows: List of dicts — one dict per row, keys are column headers.
            path: Output file path (``.csv``).
            columns: Optional column ordering. If None, columns are derived
                     from the keys of the first row's dict.

        Returns:
            The resolved output ``path``.

        Raises:
            DataExportError: S-DE-001 (write failure), S-DE-003 (dir not writable),
                             S-DE-004 (overwrite guard).
        """
        path = Path(path)

        # L16 — ensure parent directory exists and is writable
        parent = path.parent
        try:
            parent.mkdir(parents=True, exist_ok=True)
        except (OSError, PermissionError) as e:
            raise DataExportError("S-DE-003", path=parent) from e

        if not self.overwrite and path.exists():
            raise DataExportError("S-DE-004", path=path)

        # Derive columns from first row if not specified
        if columns is None and rows:
            columns = list(rows[0].keys())
        elif columns is None:
            columns = []

        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                # BOM written automatically by utf-8-sig encoding
                writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(rows)
        except (OSError, IOError) as e:
            raise DataExportError("S-DE-001", detail=str(e)) from e

        return path

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def export_to_excel(
        self,
        rows: List[Dict[str, Any]],
        path: Path,
        sheet_name: str = "Sheet1",
        columns: Optional[List[str]] = None,
    ) -> Path:
        """Write *rows* as a single-sheet ``.xlsx`` workbook.

        Args:
            rows: List of dicts — one dict per row, keys are column headers.
            path: Output file path (``.xlsx``).
            sheet_name: Worksheet name (default ``"Sheet1"``).
            columns: Optional column ordering. If None, derived from first row.

        Returns:
            The resolved output ``path``.

        Raises:
            DataExportError: S-DE-002 (write failure), S-DE-003 (dir not writable),
                             S-DE-004 (overwrite guard).
        """
        path = Path(path)

        # L16 — ensure parent directory exists and is writable
        parent = path.parent
        try:
            parent.mkdir(parents=True, exist_ok=True)
        except (OSError, PermissionError) as e:
            raise DataExportError("S-DE-003", path=parent) from e

        if not self.overwrite and path.exists():
            raise DataExportError("S-DE-004", path=path)

        # Derive columns from first row if not specified
        if columns is None and rows:
            columns = list(rows[0].keys())
        elif columns is None:
            columns = []

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Header row — bold + frozen
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
            ws.freeze_panes = "A2"

            # Data rows
            for row_idx, row in enumerate(rows, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

            # Auto-column-width
            for col_idx in range(1, len(columns) + 1):
                max_width = 0
                for row in ws.iter_rows(
                    min_col=col_idx, max_col=col_idx,
                    min_row=1, max_row=ws.max_row,
                ):
                    for cell in row:
                        if cell.value is not None:
                            max_width = max(max_width, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 50)

            wb.save(path)
        except ImportError as e:
            raise DataExportError("S-DE-002", detail=f"openpyxl not available: {e}") from e
        except (OSError, IOError) as e:
            raise DataExportError("S-DE-002", detail=str(e)) from e

        return path

    # ------------------------------------------------------------------
    # Multi-sheet Excel export
    # ------------------------------------------------------------------

    def export_multi_sheet(
        self,
        sheets: Dict[str, List[Dict[str, Any]]],
        path: Path,
    ) -> Path:
        """Write multiple sheets into a single ``.xlsx`` workbook.

        Args:
            sheets: dict mapping sheet name → list of row dicts.
            path: Output file path (``.xlsx``).

        Returns:
            The resolved output ``path``.

        Raises:
            DataExportError: S-DE-002 (write failure), S-DE-003 (dir not writable),
                             S-DE-004 (overwrite guard).
        """
        path = Path(path)

        # L16 — ensure parent directory exists and is writable
        parent = path.parent
        try:
            parent.mkdir(parents=True, exist_ok=True)
        except (OSError, PermissionError) as e:
            raise DataExportError("S-DE-003", path=parent) from e

        if not self.overwrite and path.exists():
            raise DataExportError("S-DE-004", path=path)

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            # Remove default sheet; create named sheets
            first = True
            for sname, rows in sheets.items():
                if first:
                    ws = wb.active
                    ws.title = sname
                    first = False
                else:
                    ws = wb.create_sheet(title=sname)

                columns = list(rows[0].keys()) if rows else []

                # Header row — bold + frozen
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.font = Font(bold=True)
                ws.freeze_panes = "A2"

                # Data rows
                for row_idx, row in enumerate(rows, start=2):
                    for col_idx, col_name in enumerate(columns, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

                # Auto-column-width
                for col_idx in range(1, len(columns) + 1):
                    max_width = 0
                    for row in ws.iter_rows(
                        min_col=col_idx, max_col=col_idx,
                        min_row=1, max_row=ws.max_row,
                    ):
                        for cell in row:
                            if cell.value is not None:
                                max_width = max(max_width, len(str(cell.value)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 50)

            wb.save(path)
        except ImportError as e:
            raise DataExportError("S-DE-002", detail=f"openpyxl not available: {e}") from e
        except (OSError, IOError) as e:
            raise DataExportError("S-DE-002", detail=str(e)) from e

        return path


# ---------------------------------------------------------------------------
# Module-level convenience functions
# ---------------------------------------------------------------------------

def export_to_csv(
    rows: List[Dict[str, Any]],
    path: Path,
    columns: Optional[List[str]] = None,
    overwrite: bool = True,
) -> Path:
    """Standalone CSV export — see :meth:`DataExporter.export_to_csv`."""
    return DataExporter(overwrite=overwrite).export_to_csv(rows, path, columns=columns)


def export_to_excel(
    rows: List[Dict[str, Any]],
    path: Path,
    sheet_name: str = "Sheet1",
    columns: Optional[List[str]] = None,
    overwrite: bool = True,
) -> Path:
    """Standalone Excel export — see :meth:`DataExporter.export_to_excel`."""
    return DataExporter(overwrite=overwrite).export_to_excel(
        rows, path, sheet_name=sheet_name, columns=columns,
    )


def export_multi_sheet(
    sheets: Dict[str, List[Dict[str, Any]]],
    path: Path,
    overwrite: bool = True,
) -> Path:
    """Standalone multi-sheet export — see :meth:`DataExporter.export_multi_sheet`."""
    return DataExporter(overwrite=overwrite).export_multi_sheet(sheets, path)
