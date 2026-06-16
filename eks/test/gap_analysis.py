import openpyxl
from collections import Counter

wb = openpyxl.load_workbook(
    'eks/data/twrp/datadrop/Datadrop Summary.xlsx',
    read_only=True, data_only=True
)

# --- Sheet3 fill summary ---
print("=== Sheet3: fill summary ===")
for row in wb['Sheet3'].iter_rows(min_row=1, values_only=True):
    print(row)

# --- Pipeline duplicate KEYTAG check ---
ws = wb['Pipeline']
keytags = [r[0] for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
c = Counter(keytags)
dupes = {k: v for k, v in c.items() if v > 1}
print(f"\n=== Pipeline duplicate KEYTAGs: {len(dupes)} unique tags with >1 row ===")
for k, v in list(dupes.items())[:5]:
    print(f"  {k}: {v} rows")

# --- Per-sheet: columns not mapped in appendix schema ---
SCHEMA_MAPPED = {
    'Equipment': {
        'KEYTAG','TAG_TYPE','TAG_NO','CONTRACT INFO','PID NUMBER','PROJECT PREFIX',
        'UNIT','SERVICE','DEVICE TYPE CODE','TAG LOOP NUMBER','TAG SUFFIX',
        'DESCRIPTION','SHORT DESCR','HAZARDOUS ZONE','DESIGN CAPACITY',
        'DESIGN CAPACITY UNIT','OPERATING PRESSURE - NORMAL','DUTY STAND BY',
        'MANUFACTURER 2D MODEL FILE NAME','OPERATING PRESSURE - MIN',
        'OPERATING PRESSURE - MAX','OPERATING TEMPERATURE - NORMAL','NPSH - MIN',
        'HEAD LOSS','SUCTION NOZZLE SIZE','DISCHARGE NOZZLE SIZE',
        'EQUIPMENT MATERIAL','MECH SEAL/GLAND PACKAGING','LCS TYPE','PLC PANEL',
        'PLC PANEL LOCATION','RIO PANEL','RIO PANEL LOCATION','MODEL NUMBER',
        'SERIAL NUMBER','BRAND','MANUFACTURER NAME','MANUFACTURER WEBPAGE',
        'MANUFACTURER PHONE','MANUFACTURER EMAIL','MANUFACTURER LOCATION',
        'MANUFACTURER 3D MODEL FILE NAME','LOT NUMBER','MANUFACTURE DATE','RPM',
        'EFFICIENCY_DESIGN_CAP','IMPELLER_TYPE','IMPELLER_MATERIAL',
        'ROTOR_MATERIAL','STATOR_MATERIAL','SEAL_TYPE_MECH',
        'PUB_ACE_CATEGORY','PUB_GENERIC_EQUIPMENT_TYPE','ACE ASSET CLASS',
        'PUB_LIFE_SPAN','PUB_SUPPLIER','PUB_COST_CENTER','PUB_REPLACEMENT_COST',
        'PUB_WBS_ELEMENT','PUB_EST_REPLACEMENT_DATE','PUB_DATE_OF_COMMISSION',
        'PUB_WARRANTY_TERMS','PUB_WARRANTY_START_DATE','PUB_WARRANTY_EXPIRY_DATE',
        'PUB_PRODUCT_CERTIFICATION','PUB_ACE_ASSET_NUMBER','PUB_ACE_ASSET_SUB_NUMBER',
    },
    'Motor': {
        'KEYTAG','TAG_TYPE','TAG_NO','CONTRACT INFO','P&ID DRAWING','NAME',
        'PROJECT PREFIX','UNIT','SERVICE','DEVICE TYPE CODE','TAG LOOP NUMBER',
        'TAG SUFFIX','DESCRIPTION','SHORT DESCR','EQUIPMENT NUMBER',
        'MOTOR - VOLTAGE','MOTOR - HERTZ','MOTOR - PHASE','ACTUATOR - MOTOR RATING',
        'MANUFACTURER 2D MODEL FILE NAME','HAZARDOUS ZONE',
        'OPERATING TEMPERATURE - NORMAL','MOTOR - RPM','LCS TYPE','PLC PANEL',
        'PLC PANEL LOCATION','RIO PANEL','RIO PANEL LOCATION',
        'ACTUATOR - RPM','ACTUATOR - TORQUE SETTING RANGE','ACTUATOR - RATED VOLTAGE',
        'ACTUATOR - RATED FREQUENCY','ACTUATOR - RATED CURRENT','ACTUATOR - STEM DIRECTION',
        'MODEL NUMBER','SERIAL NUMBER','BRAND','MANUFACTURER NAME','MANUFACTURER WEBPAGE',
        'MANUFACTURER PHONE','MANUFACTURER EMAIL','MANUFACTURER LOCATION',
        'MANUFACTURER 3D MODEL FILE NAME','LOT NUMBER','MANUFACTURE DATE',
        'MOTOR - RATED CURRENT','MOTOR - TORQUE SETTING RANGE',
        'ACE CATEGORY','GENERIC EQUIPMENT TYPE','ACE ASSET CLASS','LIFE SPAN',
        'SUPPLIER','COST CENTER','REPLACEMENT COST','WBS ELEMENT',
        'EST REPLACEMENT DATE','DATE OF COMMISSION','WARRANTY TERMS',
        'WARRANTY START DATE','WARRANTY EXPIRY DATE','PRODUCT CERTIFICATION',
        'ACE ASSET NUMBER','ACE ASSET SUB NUMBER',
    },
}

for sheet in ['Equipment', 'Motor']:
    ws = wb[sheet]
    headers = set(h for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) if h)
    unmapped = headers - SCHEMA_MAPPED.get(sheet, set())
    print(f"\n=== {sheet}: UNMAPPED columns ({len(unmapped)}) ===")
    for h in sorted(unmapped):
        print(f"  {h}")

# --- CONTROLVALVE: columns beyond what appendix captures ---
ws = wb['CONTROLVALVE']
headers = [h for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) if h]
print(f"\n=== CONTROLVALVE all {len(headers)} columns ===")
# highlight schema gaps: actuator ACE lifecycle block + IINT1/A_ACT_DISPLAY/FAIL/LOCKED
gaps = [h for h in headers if any(x in str(h).upper() for x in
        ['IINT','A_ACT_DISPLAY','FAIL MODE','LOCKED','- ACTUATOR','ACTUATOR - LOCKED',
         'ACTUATOR - RATED TORQUE'])]
print("  Schema gap candidates:", gaps)

# --- Instrument: extra columns not in appendix ---
ws = wb['Instrument']
headers = [h for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) if h]
extra = [h for h in headers if any(x in str(h).upper() for x in
        ['THERMOWELL','FUNCTION_CONTACT','HOUSING','LIGHTNING','SAMPLE POINT',
         'OPERATING RANGE','OUTPUT SIGNAL 2','AMS','LOLO_','LO_ALARM_TP',
         'HI_ALARM_TP','HIHI_'])]
print(f"\n=== Instrument extra/unlisted columns ===")
for h in extra:
    print(f"  {h}")

# --- MANUALVALVE: LOCKED POSITION column ---
ws = wb['MANUALVALVE']
headers = [h for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) if h]
print(f"\n=== MANUALVALVE columns not in appendix ===")
known = {'KEYTAG','TAG_TYPE','TAG_NO','CONTRACT INFO','DOC_FNAME','PROJECT PREFIX',
         'UNIT','SERVICE','DEVICE TYPE CODE','TAG LOOP NUMBER','TAG SUFFIX',
         'DESCRIPTION','SHORT DESCR','HAZARDOUS ZONE','VALVE - DUTY',
         'PIPELINE TAG NUMBER','DESIGN PRESSURE','PIPE SIZE - NOMINAL (MM)',
         'PRESSURE RATING','MANUFACTURER 2D MODEL FILE NAME',
         'FLOW RATE - NOMINAL','FLOW RATE - MIN','FLOW RATE - MAX',
         'OPERATING TEMPERATURE - NORMAL','OPERATING PRESSURE - NORMAL',
         'OPERATING PRESSURE - MIN','OPERATING PRESSURE - MAX','TEST PRESSURE',
         'BODY MATERIAL','STEM MATERIAL','VALVE CLOSURE ELEMENT','SEAT MATERIAL',
         'LINING MATERIAL','END CONDITION','MODEL NUMBER','SERIAL NUMBER','BRAND',
         'MANUFACTURER NAME','MANUFACTURER WEBPAGE','MANUFACTURER PHONE',
         'MANUFACTURER EMAIL','MANUFACTURER LOCATION','MANUFACTURER 3D MODEL FILE NAME',
         'LOT NUMBER','MANUFACTURE DATE','ACE CATEGORY','GENERIC EQUIPMENT TYPE',
         'ACE ASSET CLASS','LIFE SPAN','SUPPLIER','COST CENTER','REPLACEMENT COST',
         'WBS ELEMENT','EST REPLACEMENT DATE','DATE OF COMMISSION','WARRANTY TERMS',
         'WARRANTY START DATE','WARRANTY EXPIRY DATE','PRODUCT CERTIFICATION',
         'ACE ASSET NUMBER','ACE ASSET SUB NUMBER'}
for h in headers:
    if h not in known:
        print(f"  {h}")

print("\nDONE")
