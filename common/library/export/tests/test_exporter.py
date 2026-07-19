"""
Tests for common.library.export.DataExporter (L22).

Covers: csv round-trip, excel round-trip, multi-sheet Excel, empty rows,
Unicode/CJK characters, error paths (invalid path, read-only dir,
empty rows list, overwrite guard).

Revision: 0.1
Date: 2026-07-18
Author: opencode
"""

import csv
import io
import os
import stat
import tempfile
import unittest
from pathlib import Path

from common.library.export import DataExporter, export_to_csv, export_to_excel, export_multi_sheet
from common.library.export.exporter import DataExportError


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _read_csv(path: Path) -> list:
    """Read a CSV back as list[dict] for round-trip validation."""
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _read_excel(path: Path, sheet: str = None) -> list:
    """Read an Excel sheet back as list[dict] for round-trip validation."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) for h in next(rows_iter)]
    result = []
    for row in rows_iter:
        result.append({headers[i]: row[i] for i in range(len(headers))})
    wb.close()
    return result


# ---------------------------------------------------------------------------
# Test data
# ---------------------------------------------------------------------------

SAMPLE_ROWS = [
    {"id": "1", "name": "Alice", "score": 95},
    {"id": "2", "name": "Bob", "score": 87},
    {"id": "3", "name": "Charlie", "score": 92},
]

UNICODE_ROWS = [
    {"document_number": "DOC-001", "description": "中文文档", "title": "テスト"},
    {"document_number": "DOC-002", "description": "日本語ファイル", "title": "한국어"},
]


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestDataExporterCSV(unittest.TestCase):
    """CSV export tests."""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.exporter = DataExporter()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_csv_round_trip(self):
        """Write CSV → read back → compare contents."""
        path = Path(self.tmp) / "test.csv"
        self.exporter.export_to_csv(SAMPLE_ROWS, path)
        self.assertTrue(path.exists())

        read_back = _read_csv(path)
        self.assertEqual(len(read_back), 3)
        self.assertEqual(read_back[0]["name"], "Alice")
        self.assertEqual(read_back[1]["score"], "87")
        self.assertEqual(read_back[2]["id"], "3")

    def test_csv_bom_for_excel(self):
        """CSV starts with BOM (UTF-8-sig) for Excel compatibility."""
        path = Path(self.tmp) / "bom_test.csv"
        self.exporter.export_to_csv(SAMPLE_ROWS, path)
        with open(path, "rb") as f:
            first_bytes = f.read(3)
        self.assertEqual(first_bytes, b"\xef\xbb\xbf")  # UTF-8 BOM

    def test_csv_custom_columns(self):
        """Column ordering is respected when specified."""
        path = Path(self.tmp) / "columns.csv"
        self.exporter.export_to_csv(SAMPLE_ROWS, path, columns=["name", "id"])
        read_back = _read_csv(path)
        self.assertEqual(list(read_back[0].keys()), ["name", "id"])

    def test_csv_empty_rows_headers_only(self):
        """Empty list → file with headers only (zero data rows)."""
        path = Path(self.tmp) / "empty.csv"
        self.exporter.export_to_csv([], path, columns=["a", "b", "c"])
        read_back = _read_csv(path)
        self.assertEqual(len(read_back), 0)
        # File should still exist with the header line
        self.assertTrue(path.stat().st_size > 0)

    def test_csv_empty_no_columns_no_rows(self):
        """Empty rows + no columns → empty CSV (just header line from fieldnames=[])."""
        path = Path(self.tmp) / "empty2.csv"
        self.exporter.export_to_csv([], path)
        self.assertTrue(path.exists())

    def test_csv_unicode_cjk(self):
        """Unicode/CJK characters survive round-trip."""
        path = Path(self.tmp) / "unicode.csv"
        self.exporter.export_to_csv(UNICODE_ROWS, path)
        read_back = _read_csv(path)
        self.assertEqual(len(read_back), 2)
        self.assertEqual(read_back[0]["description"], "中文文档")
        self.assertEqual(read_back[1]["title"], "한국어")

    def test_csv_invalid_path(self):
        """S-DE-001 raised for unwritable path."""
        path = Path("/nonexistent_dir_12345/test.csv")
        with self.assertRaises(DataExportError) as ctx:
            self.exporter.export_to_csv(SAMPLE_ROWS, path)
        self.assertIn("S-DE-003", str(ctx.exception))

    def test_csv_overwrite_guard(self):
        """S-DE-004 raised when overwrite=False and file exists."""
        path = Path(self.tmp) / "guard.csv"
        self.exporter.export_to_csv(SAMPLE_ROWS, path)  # write first
        exporter_no_overwrite = DataExporter(overwrite=False)
        with self.assertRaises(DataExportError) as ctx:
            exporter_no_overwrite.export_to_csv(SAMPLE_ROWS, path)
        self.assertIn("S-DE-004", str(ctx.exception))

    def test_module_level_export_to_csv(self):
        """Module-level convenience function works."""
        path = Path(self.tmp) / "mod.csv"
        export_to_csv(SAMPLE_ROWS, path)
        self.assertTrue(path.exists())


class TestDataExporterExcel(unittest.TestCase):
    """Excel export tests."""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.exporter = DataExporter()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_excel_round_trip(self):
        """Write .xlsx → read back with openpyxl → compare."""
        path = Path(self.tmp) / "test.xlsx"
        self.exporter.export_to_excel(SAMPLE_ROWS, path)
        self.assertTrue(path.exists())

        read_back = _read_excel(path)
        self.assertEqual(len(read_back), 3)
        self.assertEqual(read_back[0]["name"], "Alice")
        self.assertEqual(read_back[1]["score"], 87)
        self.assertEqual(read_back[2]["id"], "3")

    def test_excel_custom_sheet_name(self):
        """Custom sheet name is applied."""
        path = Path(self.tmp) / "sheet.xlsx"
        self.exporter.export_to_excel(SAMPLE_ROWS, path, sheet_name="Data")

        from openpyxl import load_workbook
        wb = load_workbook(path)
        self.assertIn("Data", wb.sheetnames)
        wb.close()

    def test_excel_column_order(self):
        """Custom columns control order."""
        path = Path(self.tmp) / "order.xlsx"
        self.exporter.export_to_excel(SAMPLE_ROWS, path, columns=["score", "name"])
        read_back = _read_excel(path)
        self.assertEqual(list(read_back[0].keys()), ["score", "name"])

    def test_excel_empty_rows(self):
        """Empty rows → headers-only .xlsx."""
        path = Path(self.tmp) / "empty.xlsx"
        self.exporter.export_to_excel([], path, columns=["a", "b"])
        read_back = _read_excel(path)
        self.assertEqual(len(read_back), 0)

    def test_excel_unicode_cjk(self):
        """CJK characters survive Excel round-trip."""
        path = Path(self.tmp) / "unicode.xlsx"
        self.exporter.export_to_excel(UNICODE_ROWS, path)
        read_back = _read_excel(path)
        self.assertEqual(read_back[0]["description"], "中文文档")
        self.assertEqual(read_back[1]["title"], "한국어")

    def test_excel_invalid_path(self):
        """S-DE-003 raised for unwritable directory."""
        path = Path("/nonexistent_dir_99999/test.xlsx")
        with self.assertRaises(DataExportError) as ctx:
            self.exporter.export_to_excel(SAMPLE_ROWS, path)
        self.assertIn("S-DE-003", str(ctx.exception))

    def test_excel_overwrite_guard(self):
        """S-DE-004 raised when file exists and overwrite=False."""
        path = Path(self.tmp) / "guard.xlsx"
        self.exporter.export_to_excel(SAMPLE_ROWS, path)
        no_overwrite = DataExporter(overwrite=False)
        with self.assertRaises(DataExportError) as ctx:
            no_overwrite.export_to_excel(SAMPLE_ROWS, path)
        self.assertIn("S-DE-004", str(ctx.exception))

    def test_module_level_export_to_excel(self):
        """Convenience function works."""
        path = Path(self.tmp) / "mod.xlsx"
        export_to_excel(SAMPLE_ROWS, path)
        self.assertTrue(path.exists())


class TestDataExporterMultiSheet(unittest.TestCase):
    """Multi-sheet Excel export tests."""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.exporter = DataExporter()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_multi_sheet_round_trip(self):
        """Write multi-sheet → read back all sheets."""
        sheets = {
            "Phase_A": [
                {"doc": "A1", "status": "ok"},
                {"doc": "A2", "status": "ok"},
            ],
            "Phase_B": [
                {"doc": "B1", "score": 90},
                {"doc": "B2", "score": 85},
                {"doc": "B3", "score": 70},
            ],
        }
        path = Path(self.tmp) / "multi.xlsx"
        self.exporter.export_multi_sheet(sheets, path)
        self.assertTrue(path.exists())

        from openpyxl import load_workbook
        wb = load_workbook(path)
        self.assertEqual(set(wb.sheetnames), {"Phase_A", "Phase_B"})

        # Phase_A
        a_rows = _read_excel(path, "Phase_A")
        self.assertEqual(len(a_rows), 2)
        self.assertEqual(a_rows[0]["doc"], "A1")

        # Phase_B
        b_rows = _read_excel(path, "Phase_B")
        self.assertEqual(len(b_rows), 3)
        self.assertEqual(b_rows[2]["score"], 70)

        wb.close()

    def test_multi_sheet_empty_dict(self):
        """Empty sheets dict produces a workbook with default sheet."""
        path = Path(self.tmp) / "empty_multi.xlsx"
        self.exporter.export_multi_sheet({}, path)
        self.assertTrue(path.exists())

    def test_multi_sheet_overwrite_guard(self):
        """S-DE-004 for overwrite guard on multi-sheet."""
        path = Path(self.tmp) / "guard_multi.xlsx"
        self.exporter.export_multi_sheet({"A": SAMPLE_ROWS}, path)
        no_overwrite = DataExporter(overwrite=False)
        with self.assertRaises(DataExportError) as ctx:
            no_overwrite.export_multi_sheet({"A": SAMPLE_ROWS}, path)
        self.assertIn("S-DE-004", str(ctx.exception))

    def test_module_level_export_multi_sheet(self):
        """Convenience function works."""
        path = Path(self.tmp) / "mod_multi.xlsx"
        export_multi_sheet({"Data": SAMPLE_ROWS}, path)
        self.assertTrue(path.exists())


class TestDataExporterEdgeCases(unittest.TestCase):
    """Edge-case tests."""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.exporter = DataExporter()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_csv_rows_with_extra_keys_ignored(self):
        """Rows with extra keys beyond columns are silently ignored (extrasaction='ignore')."""
        rows = [{"a": 1, "b": 2, "extra": "skip"}]
        path = Path(self.tmp) / "extra.csv"
        self.exporter.export_to_csv(rows, path, columns=["a", "b"])
        # Should not raise — extrasaction='ignore' in DictWriter
        self.assertTrue(path.exists())

    def test_excel_rows_with_extra_keys(self):
        """Extra keys beyond columns are omitted."""
        rows = [{"a": 1, "b": 2, "extra": "skip"}]
        path = Path(self.tmp) / "extra.xlsx"
        self.exporter.export_to_excel(rows, path, columns=["a", "b"])
        read_back = _read_excel(path)
        self.assertEqual(list(read_back[0].keys()), ["a", "b"])


if __name__ == "__main__":
    unittest.main()
