"""
XLSX Parser for EKS - Uses openpyxl for spreadsheet data extraction.
"""
import openpyxl
from pathlib import Path
from typing import Any, Dict, List
from .base_parser import BaseParser

class XLSXParser(BaseParser):
    """
    Parses XLSX files to extract sheet data as text or structured blocks.
    """
    def parse(self) -> List[Dict[str, Any]]:
        blocks = []
        wb = openpyxl.load_workbook(str(self.file_path), data_only=True, read_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                
                # Convert rows to a simple text representation for initial ingestion
                content = "\n".join(["\t".join([str(cell) if cell is not None else "" for cell in row]) for row in rows])
                
                blocks.append({
                    "content": content,
                    "type": "sheet",
                    "metadata": {
                        "sheet_name": sheet_name,
                        "row_count": len(rows),
                        "column_count": len(rows[0]) if rows else 0
                    }
                })
            return blocks
        finally:
            wb.close()

    def extract_metadata(self) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(str(self.file_path), read_only=True)
        try:
            props = wb.properties
            return {
                "author": props.creator,
                "title": props.title,
                "subject": props.subject,
                "created": props.created.isoformat() if props.created else None,
                "modified": props.modified.isoformat() if props.modified else None,
                "sheet_count": len(wb.sheetnames)
            }
        finally:
            wb.close()
