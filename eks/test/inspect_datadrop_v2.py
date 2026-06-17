import openpyxl
import json

file_path = r'eks/data/twrp/datadrop/Datadrop Summary.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    summary = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Get headers from first row
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        
        # Check first few rows for sample data (to see if it's design or physical)
        samples = []
        for row in sheet.iter_rows(min_row=2, max_row=5):
            samples.append([cell.value for cell in row])
            
        summary[sheet_name] = {
            "headers": headers,
            "samples": samples
        }
    
    print(json.dumps(summary, indent=2))
except Exception as e:
    print(f"Error: {e}")
