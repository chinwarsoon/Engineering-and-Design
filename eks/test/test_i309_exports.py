"""
Focused I309 tests: single-workbook Excel export + schema-driven column override.

Revision: 0.1
Date: 2026-08-13
Author: opencode
Summary: T1.288/T1.289/T1.291 — DataExporter.export_to_workbook (per-sheet
         column control, export_multi_sheet delegates); ONE workbook for all
         xlsx-enabled views in the pipeline export phase; schema-driven
         workbook file name (system_parameters.export_workbook_file_name);
         review_flags gains xlsx (all 3 views Excel-eligible); column override
         validated by validate_export_column_override (S-C-S-0313);
         phase1_server schema-driven export + GET /api/v1/export_views.
"""

import json
import tempfile
from pathlib import Path

import pytest

from eks.engine.pipeline_engine.exporter import (
    resolve_export_views,
    validate_export_column_override,
)


_PROJECT_ROOT = Path(__file__).resolve().parent.parent
_CONFIG_DIR = _PROJECT_ROOT / "config"
_SCHEMA_DIR = _CONFIG_DIR / "schemas"

_DEFAULT_VIEW_IDS = ("discovery_inventory", "extraction_results", "review_flags")


def _load_json(name):
    return json.loads((_SCHEMA_DIR / name).read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# T1.289 — view config: ALL 3 default views are now Excel-eligible (Q3)
# ---------------------------------------------------------------------------

def test_i309_all_default_views_xlsx_eligible():
    """T1.289 (Q3): every default view declares xlsx in formats."""
    cfg = _load_json("eks_export_view_config.json")
    assert cfg["version"] == "1.2.0"  # I309 bump (was 1.1.0 in I308)
    for entry in cfg["views"]:
        assert entry["view_id"] in _DEFAULT_VIEW_IDS
        assert "xlsx" in entry["formats"], (
            f"view '{entry['view_id']}' must be Excel-eligible (Q3)"
        )
        assert "csv" in entry["formats"]


def test_i309_resolve_export_views_return_formats():
    """T1.289: resolve_export_views surfaces per-view formats."""
    specs = resolve_export_views(_SCHEMA_DIR)
    for view_id in _DEFAULT_VIEW_IDS:
        assert "xlsx" in specs[view_id]["formats"], view_id


# ---------------------------------------------------------------------------
# T1.289 — schema-driven workbook file name (Q6)
# ---------------------------------------------------------------------------

def test_i309_workbook_file_name_in_config():
    """T1.289 (Q6): system_parameters.export_workbook_file_name present in
    eks_config.json (schema-driven SSOT for the single-workbook export)."""
    cfg = _load_json("eks_config.json")
    assert cfg["version"] == "1.14.0"
    sp = cfg["system_parameters"]
    assert sp.get("export_workbook_file_name") == "eks_export.xlsx"


def test_i309_workbook_file_name_in_base_schema():
    """T1.289 (Q6): export_workbook_file_name declared (required) in
    system_parameters_def of eks_base_schema.json."""
    base = _load_json("eks_base_schema.json")
    assert base["version"] == "1.22.0"
    sp_def = base["definitions"]["system_parameters_def"]
    assert "export_workbook_file_name" in sp_def["properties"]
    prop = sp_def["properties"]["export_workbook_file_name"]
    assert prop.get("type") == "string"
    assert prop.get("default") == "eks_export.xlsx"
    assert "export_workbook_file_name" in sp_def["required"]
    assert sp_def["additionalProperties"] is False


# ---------------------------------------------------------------------------
# T1.289 — S-C-S-0313 EXPORT_COLUMN_NOT_ALLOWED registered (Q6 cross-source)
# ---------------------------------------------------------------------------

def test_i309_error_code_registered():
    """T1.289: S-C-S-0313 EXPORT_COLUMN_NOT_ALLOWED registered with all fields;
    metadata counts reconciled (140 = 82 system + 58 data)."""
    cfg = _load_json("eks_error_config.json")
    assert cfg["metadata"]["version"] == "1.10.0"
    assert cfg["metadata"]["total_codes"] == 140
    assert cfg["metadata"]["system_codes"] == 82
    assert cfg["metadata"]["data_logic_codes"] == 58
    entry = cfg["system_errors"]["S-C-S-0313"]
    assert entry["name"] == "EXPORT_COLUMN_NOT_ALLOWED"
    assert entry["severity"] == "HIGH"
    assert entry["category"] == "Config"
    assert entry["stops_pipeline"] is True
    assert "{column}" in entry["message"] and "{view_id}" in entry["message"]
    # config range end moved 0312 → 0313
    assert cfg["system_error_ranges"]["config"]["end_id"] == "S-C-S-0313"
    assert cfg["system_error_ranges"]["config"]["count"] == 13


def test_i309_error_code_matches_canonical_format():
    """T1.289: S-C-S-0313 matches the canonical S-{cat}-S-{id4} regex."""
    base = _load_json("eks_error_code_base.json")
    import re
    pattern = re.compile(base["definitions"]["system_error_code_format"]["pattern"])
    assert pattern.match("S-C-S-0313"), "code must match system_error_code_format regex"


# ---------------------------------------------------------------------------
# T1.289 — validate_export_column_override (S-C-S-0313)
# ---------------------------------------------------------------------------

def test_i309_override_valid_subset_reorder():
    """T1.289: valid override (subset/reorder) is normalized and returned."""
    result = validate_export_column_override(
        _SCHEMA_DIR,
        {"discovery_inventory": ["document_type", "document_number"]},
    )
    assert result == {
        "discovery_inventory": ["document_type", "document_number"],
    }


def test_i309_override_disallowed_column_raises():
    """T1.289: a column outside the view's config columns → S-C-S-0313."""
    with pytest.raises(RuntimeError) as excinfo:
        validate_export_column_override(
            _SCHEMA_DIR,
            {"discovery_inventory": ["document_number", "NOT_A_REAL_COLUMN"]},
        )
    assert "S-C-S-0313" in str(excinfo.value)
    assert "NOT_A_REAL_COLUMN" in str(excinfo.value)


def test_i309_override_unknown_view_raises():
    """T1.289: an unknown view_id → S-C-S-0313."""
    with pytest.raises(RuntimeError) as excinfo:
        validate_export_column_override(_SCHEMA_DIR, {"ghost_view": ["document_number"]})
    assert "S-C-S-0313" in str(excinfo.value)
    assert "ghost_view" in str(excinfo.value)


def test_i309_override_empty_is_noop():
    """T1.289: None/empty override → {} (no-op)."""
    assert validate_export_column_override(_SCHEMA_DIR, None) == {}
    assert validate_export_column_override(_SCHEMA_DIR, {}) == {}


def test_i309_override_missing_config_raises_0312(tmp_path):
    """T1.289: missing view config → S-C-S-0312 (no silent fallback, §16)."""
    empty_dir = tmp_path / "no_view_config"
    empty_dir.mkdir()
    with pytest.raises(RuntimeError) as excinfo:
        validate_export_column_override(empty_dir, {"discovery_inventory": ["document_number"]})
    assert "S-C-S-0312" in str(excinfo.value)


# ---------------------------------------------------------------------------
# T1.288 — DataExporter.export_to_workbook per-sheet column control (Q2)
# ---------------------------------------------------------------------------

def test_i309_export_to_workbook_per_sheet_columns():
    """T1.288: one workbook, per-sheet ordered columns (keys = sheet names)."""
    from common.library.export import DataExporter

    rows = [
        {"document_number": "D1", "document_type": "TS", "extra": "x"},
        {"document_number": "D2", "document_type": "DR", "extra": "y"},
    ]
    sheets = {"Review Flags": rows, "Discovery Inventory": rows}
    columns = {
        "Review Flags": ["document_number", "document_type"],
    }  # Discovery Inventory derives its own
    with tempfile.TemporaryDirectory() as directory:
        path = Path(directory) / "eks_export.xlsx"
        DataExporter().export_to_workbook(sheets, path, columns=columns)
        assert path.exists()

        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        try:
            assert set(wb.sheetnames) == {"Review Flags", "Discovery Inventory"}
            ws = wb["Review Flags"]
            headers = [c.value for c in next(ws.iter_rows())]
            assert headers == ["document_number", "document_type"], headers
        finally:
            wb.close()


def test_i309_pipeline_export_no_per_view_xlsx_literals():
    """T1.289: pipeline export path must NOT emit per-view .xlsx (single
    workbook replaces per-view export_to_excel for xlsx)."""
    pipeline_path = _PROJECT_ROOT / "engine" / "eks_engine_pipeline.py"
    src = pipeline_path.read_text(encoding="utf-8")
    # xlsx now flows through one export_to_workbook call, not export_to_excel
    assert 'export_to_workbook(' in src
    # per-view xlsx filename literal must be gone (only .csv per-view remains)
    assert "{base_name}.xlsx" not in src


# ---------------------------------------------------------------------------
# T1.290 — phase1_server schema-driven export + export_views endpoint
# ---------------------------------------------------------------------------

def test_i309_phase1_server_no_hardcoded_export_columns():
    """T1.290 (Q1): hardcoded export column lists removed from phase1_server."""
    server_path = _PROJECT_ROOT / "ui" / "backend" / "phase1_server.py"
    src = server_path.read_text(encoding="utf-8")
    assert "resolve_export_views" in src
    assert "validate_export_column_override" in src
    # Old hardcoded 6-column discovery phase_defs array is gone (column lists
    # now come from eks_export_view_config.json via resolve_export_views).
    assert "file_path\", \"ingested_at\"" not in src
    # New endpoints/methods present
    assert '"export_views"' in src  # GET /api/v1/export_views route
    assert "def _handle_export_views" in src
    assert "def _mk_rows_fn" in src


def test_i309_phase1_export_views_endpoint_declared():
    """T1.290: GET /api/v1/export_views route wired into do_GET."""
    server_path = _PROJECT_ROOT / "ui" / "backend" / "phase1_server.py"
    src = server_path.read_text(encoding="utf-8")
    assert '["api", "v1", "export_views"]' in src


def test_i309_phase1_run_post_accepts_export_columns():
    """T1.290 (Q4): run POST wires + validates export_columns override."""
    server_path = _PROJECT_ROOT / "ui" / "backend" / "phase1_server.py"
    src = server_path.read_text(encoding="utf-8")
    assert "export_columns" in src
    assert "validate_export_column_override" in src
    assert "_export_columns_override" in src


def test_i314_phase1_server_schema_driven_download_file_name():
    """I314 follow-up: download endpoint file names are schema-driven. The
    hardcoded 'eks_export_phase_{phase}.{fmt}' literal is removed and the
    template field is registered in both schema layers (SSOT)."""
    server_path = _PROJECT_ROOT / "ui" / "backend" / "phase1_server.py"
    src = server_path.read_text(encoding="utf-8")
    # schema-driven template read at runtime
    assert "export_download_file_name_template" in src
    # legacy hardcoded per-phase download name is gone
    assert "eks_export_phase_" not in src
    # SSOT: field present in base schema definition + default config values
    base_src = (_SCHEMA_DIR / "eks_base_schema.json").read_text(encoding="utf-8")
    cfg_src = (_SCHEMA_DIR / "eks_config.json").read_text(encoding="utf-8")
    assert "export_download_file_name_template" in base_src
    assert "export_download_file_name_template" in cfg_src


# ---------------------------------------------------------------------------
# T1.288 — shared exporter package exports export_to_workbook
# ---------------------------------------------------------------------------

def test_i309_export_package_reexports_workbook():
    """T1.288: common.library.export re-exports export_to_workbook."""
    import common.library.export as pkg
    assert hasattr(pkg, "export_to_workbook")
    assert "export_to_workbook" in pkg.__all__


def test_i309_eks_engine_pipeline_reexports_helper():
    """T1.289: eks_engine_pipeline re-exports the override validator."""
    from eks.engine.eks_engine_pipeline import validate_export_column_override as v
    assert v is validate_export_column_override