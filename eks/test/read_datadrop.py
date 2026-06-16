import openpyxl

wb = openpyxl.load_workbook(
    'eks/data/twrp/datadrop/Datadrop Summary.xlsx',
    read_only=True, data_only=True
)
print('Sheets:', wb.sheetnames)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    all_rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
    headers = all_rows[0] if all_rows else []
    print(f'\n=== {sheet} (max_row={ws.max_row}, max_col={ws.max_column}) ===')
    print('HEADERS:')
    for i, h in enumerate(headers):
        print(f'  [{i}] {h}')
    for r, row in enumerate(all_rows[1:], 1):
        print(f'ROW{r}: {row}')
