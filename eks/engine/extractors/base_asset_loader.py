"""BaseAssetLoader — schema-driven asset loader for the Datadrop workbook.

Tasks T1.309 / T1.311 / T1.312 (I318, I020, I021) — Phase 1 spike that proves the
I/O pattern before Phase 3 reuse (T3.9):

    Excel  ->  normalize  ->  compose  ->  validate

* T1.309  read 1-2 Datadrop sheets by ``tag_type``; apply ``column_normalization``
          from ``eks_asset_config.json``; compose base-schema fragments including
          ``conditional_fragments``; emit canonical asset records.
* T1.311  7-sheet column-coverage pre-check vs ``column_normalization`` (I020).
* T1.312  null/blank field tolerance + per-asset completeness/health score (I021);
          no silent column drops — unmapped columns are reported.

Design notes (spike scope, Phase 3 formalisation):
* Identity columns (KEYTAG, TAG_TYPE, TAG_NO, ...) are universal extractor
  columns that ``column_normalization`` does not yet carry. ``_NATIVE_ITEM_CORE_COLUMNS``
  is the documented native default used as fallback; config entries always win.
  Promotion of this map into ``column_normalization`` is planned for Phase 3.
* Paths resolve schema-first: ``default_base_path`` + ``resolve_paths`` against
  ``global_paths`` (AGENTS.md §15). Only the ``twrp/datadrop`` sub-route is a
  module constant pending Phase 3 config promotion.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from common.library.paths import resolve_paths
from common.library.paths.root_discovery import default_base_path
from eks.engine.core.schema_loader import SchemaLoader
from eks.engine.logging.logger import EKSLogger

# ---------------------------------------------------------------------------
# Native identity column map (fallback only — config ``column_normalization``
# entries always take precedence).
# ---------------------------------------------------------------------------
_NATIVE_ITEM_CORE_COLUMNS: Dict[str, str] = {
    "KEYTAG": "keytag",
    "TAG_TYPE": "tag_type",
    "TAG_NO": "tag_no",
    "CONTRACT INFO": "contract_info",
    "PROJECT PREFIX": "project_prefix",
    "SERVICE": "service",
    "DEVICE TYPE CODE": "device_type_code",
    "TAG LOOP NUMBER": "tag_loop_number",
    "TAG SUFFIX": "tag_suffix",
    "HAZARDOUS ZONE": "hazardous_zone",
    "UNIT": "unit",
    "DESCRIPTION": "description",
    "SHORT DESCR": "short_description",
    "PID NUMBER": "p_and_id_file",
}

# Relative location of the Datadrop workbook under the resolved data_dir.
# Pending Phase 3 config promotion (schema-driven SSOT).
_DATADROP_SUBPATH = ("twrp", "datadrop", "Datadrop Summary.xlsx")

# Health grade thresholds (T1.312).
_HEALTH_HIGH = 0.7
_HEALTH_MEDIUM = 0.4


@dataclass
class SheetData:
    """Raw sheet content: headers plus one dict per data row (None for blanks)."""

    sheet_name: str
    headers: List[str] = field(default_factory=list)
    rows: List[Dict[str, Any]] = field(default_factory=list)


@dataclass
class AssetLoadResult:
    """Result of a ``load_sheet`` call."""

    sheet_name: str
    tag_type: Optional[str]
    records: List[Dict[str, Any]] = field(default_factory=list)
    skipped_rows: List[Dict[str, Any]] = field(default_factory=list)
    unmapped_columns: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


class BaseAssetLoader:
    """Loads canonical asset records from the Datadrop workbook.

    Args:
        config_dir: EKS config directory holding ``eks_asset_config.json`` and
            the asset schemas (default: anchor-discovered ``<repo>/eks/config``).
        workbook_path: Optional explicit path to the Datadrop workbook.
            Default: resolved via ``global_paths`` + ``_DATADROP_SUBPATH``.
        logger: Optional :class:`EKSLogger`; a module logger is created otherwise.
    """

    def __init__(
        self,
        config_dir: Optional[str | Path] = None,
        workbook_path: Optional[str | Path] = None,
        logger: Optional[EKSLogger] = None,
    ) -> None:
        repo_root = default_base_path("eks", reference=__file__)
        self._repo_root = repo_root
        self._config_dir = Path(config_dir) if config_dir else repo_root / "eks" / "config"
        self._schema_loader = SchemaLoader(config_dir=self._config_dir)
        self._schema_loader.load_all()

        self.config: Dict[str, Any] = self._schema_loader.asset_config
        self.base_schema: Dict[str, Any] = self._schema_loader.asset_base_schema
        self.logger: EKSLogger = logger or EKSLogger("BaseAssetLoader")

        self.workbook_path: Path = (
            Path(workbook_path) if workbook_path else self._default_workbook_path()
        )
        self._workbook = None
        self._fragment_fields: Optional[Dict[str, Dict[str, Any]]] = None

    # ------------------------------------------------------------------ path
    def _default_workbook_path(self) -> Path:
        """Resolve the Datadrop workbook via schema-driven global_paths (SSOT)."""
        resolved = resolve_paths(self._repo_root, self._schema_loader.config).resolve(
            self._repo_root
        )
        return Path(resolved["data_dir"]).joinpath(*_DATADROP_SUBPATH)

    # ------------------------------------------------------------ workbook
    def _open_workbook(self):
        """Open the workbook lazily (data_only: formulas resolved to values)."""
        if self._workbook is None:
            if not self.workbook_path.is_file():
                raise FileNotFoundError(
                    f"Datadrop workbook not found: {self.workbook_path} "
                    f"(resolved from global_paths + {_DATADROP_SUBPATH})."
                )
            self.logger.status(f"Opening workbook: {self.workbook_path.name}")
            self._workbook = load_workbook(self.workbook_path, data_only=True, read_only=True)
        return self._workbook

    def close(self) -> None:
        """Release the openpyxl workbook handle."""
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None

    @property
    def sheet_names(self) -> List[str]:
        return list(self._open_workbook().sheetnames)

    def _read_headers(self, sheet_name: str) -> List[str]:
        wb = self._open_workbook()
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"sheet {sheet_name!r} not in workbook {self.workbook_path.name}; "
                f"available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        first = next(ws.iter_rows(values_only=True), None)
        if not first:
            return []
        return [str(h).strip() if h is not None else "" for h in first]

    def _read_sheet(self, sheet_name: str) -> SheetData:
        """Read a sheet into :class:`SheetData`; blank cells become None."""
        wb = self._open_workbook()
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"sheet {sheet_name!r} not in workbook {self.workbook_path.name}; "
                f"available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return SheetData(sheet_name)
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows: List[Dict[str, Any]] = []
        for raw in rows[1:]:
            row: Dict[str, Any] = {}
            populated = False
            for i, header in enumerate(headers):
                if not header:
                    continue
                value = raw[i] if i < len(raw) else None
                if isinstance(value, str):
                    value = value.strip() or None
                if value is not None:
                    populated = True
                row[header] = value
            if populated:
                data_rows.append(row)
        return SheetData(sheet_name, headers, data_rows)

    # ------------------------------------------------------------ fragments
    @property
    def fragment_fields(self) -> Dict[str, Dict[str, Any]]:
        """canonical fragment -> {field: None | {"$nested": {child, ...}}} from base schema."""
        if self._fragment_fields is None:
            self._fragment_fields = self._build_fragment_field_map()
        return self._fragment_fields

    def _build_fragment_field_map(self) -> Dict[str, Dict[str, Any]]:
        definitions = self.base_schema.get("definitions", {})
        fragment_map: Dict[str, Dict[str, Any]] = {}
        for name, frag in definitions.items():
            if not isinstance(frag, dict):
                continue
            props = frag.get("properties", {})
            if not isinstance(props, dict):
                continue
            fields: Dict[str, Any] = {}
            for pname, pspec in props.items():
                if (
                    isinstance(pspec, dict)
                    and pspec.get("type") == "object"
                    and isinstance(pspec.get("properties"), dict)
                    and pspec["properties"]
                ):
                    fields[pname] = {"$nested": set(pspec["properties"].keys())}
                else:
                    fields[pname] = None
            fragment_map[name] = fields
        return fragment_map

    # ------------------------------------------------------------- normalize
    def _column_target(self, sheet_name: str, header: str) -> Optional[str]:
        """Config ``column_normalization`` wins; native item_core map is fallback."""
        config_map = self.config.get("column_normalization", {}).get(sheet_name, {})
        return config_map.get(header) or _NATIVE_ITEM_CORE_COLUMNS.get(header)

    def _normalize_row(
        self, sheet_name: str, raw: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Map raw sheet row to canonical field names; null/blank values are skipped."""
        normalized: Dict[str, Any] = {}
        for header, value in raw.items():
            target = self._column_target(sheet_name, header)
            if target is None or value is None:
                continue
            normalized[target] = value
        return normalized

    # -------------------------------------------------------------- compose
    def _compose_record(
        self, normalized: Dict[str, Any], source_meta: Dict[str, Any]
    ) -> "tuple[Optional[Dict[str, Any]], List[str]]":
        """Compose a canonical record for one normalized row.

        Returns ``(record, unrouted_fields)``; ``record`` is None when the row's
        tag_type is absent from ``asset_type_registry`` (row is then skipped).
        """
        tag_type = normalized.get("tag_type") or source_meta.get("tag_type")
        registry = self.config.get("asset_type_registry", {})
        entry = registry.get(tag_type) if isinstance(tag_type, str) else None
        if not entry:
            return None, [f"tag_type {tag_type!r} not in asset_type_registry"]

        fragments = list(entry.get("fragments", []) or [])
        for cond in entry.get("conditional_fragments", []) or []:
            when_value = normalized.get(cond.get("when"))
            allowed = cond.get("in") or []
            if isinstance(when_value, str) and when_value in allowed:
                fragments.append(cond.get("fragment"))

        record: Dict[str, Any] = {"source_meta": source_meta}
        for frag in fragments:
            record[frag] = {}
        # Pre-materialise nested objects (e.g. asset_context's hierarchy
        # objects) so every canonical record has a stable, schema-shaped key set.
        for frag in fragments:
            for field, spec in self.fragment_fields.get(frag, {}).items():
                if isinstance(spec, dict) and spec.get("$nested"):
                    record[frag][field] = {}
        unrouted = self._assign_fields(record, normalized, fragments)
        record["asset_health"] = self._compute_health(record, fragments, normalized)
        return record, unrouted

    def _assign_fields(
        self,
        record: Dict[str, Any],
        normalized: Dict[str, Any],
        fragments: List[str],
    ) -> List[str]:
        """Route canonical fields into their declaring fragments (dotted-aware).

        Dotted targets (``system_hierarchy.system_code``) land in nested objects
        under ``asset_context`` (or any fragment whose nested object declares the
        child field). Fields that no composed fragment declares are returned as
        ``unrouted`` — never silently dropped (I021).
        """
        fragment_map = self.fragment_fields
        unrouted: List[str] = []
        for canon, value in normalized.items():
            if value is None or (isinstance(value, str) and not value.strip()):
                # Null/blank tolerance (T1.312): blank fields are never written
                # into the canonical record; they surface via asset_health.
                continue
            if "." in canon:
                parent, child = canon.split(".", 1)
                routed = False
                for frag in fragments:
                    spec = fragment_map.get(frag, {}).get(parent)
                    if isinstance(spec, dict) and child in spec.get("$nested", set()):
                        record[frag].setdefault(parent, {})[child] = value
                        routed = True
                        break
                if not routed:
                    unrouted.append(canon)
                continue
            routed = False
            for frag in fragments:
                if canon in fragment_map.get(frag, {}):
                    record[frag][canon] = value
                    routed = True
                    break
            if not routed:
                unrouted.append(canon)
        return unrouted

    # --------------------------------------------------------------- health
    def _compute_health(
        self,
        record: Dict[str, Any],
        fragments: List[str],
        normalized: Dict[str, Any],
    ) -> Dict[str, Any]:
        """Per-asset completeness/health (T1.312).

        * ``completeness`` — populated / declared canonical fields across the
          composed fragments (data quality vs the asset type's schema).
        * ``fill_rate``    — populated / fields the sheet made available for
          this asset type (row-level fill given the source).
        * ``grade``        — HIGH / MEDIUM / LOW from ``completeness``.
        * ``null_fields``  — canonical paths that were blank on this row.
        """
        fragment_map = self.fragment_fields
        frag_stats: Dict[str, Any] = {}
        total_known = 0
        total_filled = 0
        null_fields: List[str] = []
        for frag in fragments:
            fields = fragment_map.get(frag, {})
            known = 0
            filled = 0
            for canon, spec in fields.items():
                if isinstance(spec, dict):
                    child_obj = record.get(frag, {}).get(canon, {})
                    for child in spec["$nested"]:
                        known += 1
                        if child_obj.get(child) not in (None, ""):
                            filled += 1
                        else:
                            null_fields.append(f"{frag}.{canon}.{child}")
                else:
                    known += 1
                    if record.get(frag, {}).get(canon) not in (None, ""):
                        filled += 1
                    else:
                        null_fields.append(f"{frag}.{canon}")
            frag_stats[frag] = {
                "known": known,
                "filled": filled,
                "ratio": round(filled / known, 3) if known else 1.0,
            }
            total_known += known
            total_filled += filled
        completeness = round(total_filled / total_known, 3) if total_known else 1.0

        available = len(normalized)
        fill_rate = round(total_filled / available, 3) if available else 1.0
        grade = (
            "HIGH"
            if completeness >= _HEALTH_HIGH
            else ("MEDIUM" if completeness >= _HEALTH_MEDIUM else "LOW")
        )
        return {
            "completeness": completeness,
            "fill_rate": fill_rate,
            "grade": grade,
            "fragment_completeness": frag_stats,
            "null_fields": null_fields,
        }

    # ----------------------------------------------------------------- load
    def load_sheet(
        self,
        sheet_name: str,
        tag_type: Optional[str] = None,
        limit: Optional[int] = None,
    ) -> AssetLoadResult:
        """Load one sheet into canonical asset records.

        Args:
            sheet_name: Sheet to read (see ``sheet_names``).
            tag_type: Optional AT_* filter. When given, only rows whose TAG_TYPE
                matches are loaded and unknown values are skipped. When omitted,
                each row's TAG_TYPE drives composition.
            limit: Cap on the number of records returned (spike convenience).
        """
        sheet = self._read_sheet(sheet_name)
        mapped_columns = set(self.config.get("column_normalization", {}).get(sheet_name, {}))
        mapped_columns |= set(_NATIVE_ITEM_CORE_COLUMNS)
        unmapped = sorted(
            {h for h in sheet.headers if h and h not in mapped_columns}
        )
        warnings: List[str] = []
        if unmapped:
            warnings.append(
                f"sheet {sheet_name!r}: {len(unmapped)} unmapped column(s): {unmapped}"
            )

        records: List[Dict[str, Any]] = []
        skipped: List[Dict[str, Any]] = []
        for idx, raw in enumerate(sheet.rows, start=2):
            if limit is not None and len(records) >= limit:
                break
            normalized = self._normalize_row(sheet_name, raw)
            row_tag_type = normalized.get("tag_type") or tag_type
            if tag_type and row_tag_type != tag_type:
                continue
            if not isinstance(row_tag_type, str):
                skipped.append({"row": idx, "reason": "missing tag_type"})
                continue
            source_meta = {
                "source": self.workbook_path.name,
                "sheet": sheet_name,
                "row": idx,
                "tag_type": row_tag_type,
            }
            record, unrouted = self._compose_record(normalized, source_meta)
            if record is None:
                skipped.append({"row": idx, "reason": unrouted[0] if unrouted else "unknown"})
                continue
            if unrouted:
                warnings.append(
                    f"row {idx}: {len(unrouted)} field(s) not declared by composed "
                    f"fragments: {unrouted}"
                )
            records.append(record)

        result = AssetLoadResult(
            sheet_name=sheet_name,
            tag_type=tag_type,
            records=records,
            skipped_rows=skipped,
            unmapped_columns=unmapped,
            warnings=warnings,
        )
        self.logger.status(
            f"Loaded {len(records)} asset(s) from {sheet_name!r} "
            f"(tag_type={tag_type!r}, skipped={len(skipped)})"
        )
        return result

    # ------------------------------------------------------------ coverage
    def column_coverage(self, sheet_names: Optional[List[str]] = None) -> Dict[str, Dict[str, Any]]:
        """T1.311 — diff every column of the given sheets against the config map.

        Returns per sheet:
            mapped    [(header, canonical), ...]  — via ``column_normalization``
            native    [(header, canonical), ...]  — via ``_NATIVE_ITEM_CORE_COLUMNS``
            unmapped  [header, ...]
        """
        sheet_names = list(
            sheet_names or self.config.get("column_normalization", {}).keys()
        )
        coverage: Dict[str, Dict[str, Any]] = {}
        for sheet_name in sheet_names:
            headers = self._read_headers(sheet_name)
            config_map = self.config.get("column_normalization", {}).get(sheet_name, {})
            mapped: List[tuple] = []
            native: List[tuple] = []
            unmapped: List[str] = []
            for header in headers:
                if not header:
                    continue
                if header in config_map:
                    mapped.append((header, config_map[header]))
                elif header in _NATIVE_ITEM_CORE_COLUMNS:
                    native.append((header, _NATIVE_ITEM_CORE_COLUMNS[header]))
                else:
                    unmapped.append(header)
            coverage[sheet_name] = {
                "mapped": mapped,
                "native": native,
                "unmapped": unmapped,
            }
        return coverage

    def format_coverage_report(self, coverage: Dict[str, Dict[str, Any]]) -> str:
        """Render the T1.311 column-coverage report as markdown."""
        lines = [
            "# Asset Column-Coverage Pre-check (T1.311 / I020)",
            "",
            f"- Workbook: `{self.workbook_path}`",
            f"- Config map source: `column_normalization` + native item_core identity map",
            f"- Sheets analysed: {len(coverage)}",
            "",
            "| Sheet | Config-mapped | Native-mapped | Unmapped |",
            "|-------|--------------:|--------------:|---------:|",
        ]
        for sheet_name in sorted(coverage):
            c = coverage[sheet_name]
            lines.append(
                f"| {sheet_name} | {len(c['mapped'])} | {len(c['native'])} | "
                f"{len(c['unmapped'])} |"
            )
        lines.append("")
        for sheet_name in sorted(coverage):
            c = coverage[sheet_name]
            lines.append(f"## {sheet_name}")
            lines.append("")
            if c["mapped"]:
                lines.append("| Datadrop column | Canonical field |")
                lines.append("|-----------------|-----------------|")
                for header, canon in sorted(c["mapped"]):
                    lines.append(f"| {header} | `{canon}` |")
            else:
                lines.append("_No config-mapped columns._")
            lines.append("")
            if c["native"]:
                lines.append("Native identity mapping (not yet in `column_normalization`):")
                lines.append("")
                lines.append("| Datadrop column | Canonical field |")
                lines.append("|-----------------|-----------------|")
                for header, canon in sorted(c["native"]):
                    lines.append(f"| {header} | `{canon}` |")
                lines.append("")
            if c["unmapped"]:
                lines.append(f"Unmapped columns ({len(c['unmapped'])}):")
                lines.append("")
                for header in sorted(c["unmapped"]):
                    lines.append(f"- `{header}`")
            else:
                lines.append("_No unmapped columns._")
            lines.append("")
        return "\n".join(lines)

    def write_coverage_report(
        self, output_path: str | Path, sheet_names: Optional[List[str]] = None
    ) -> Path:
        """Generate the T1.311 markdown report to *output_path*."""
        coverage = self.column_coverage(sheet_names)
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(self.format_coverage_report(coverage), encoding="utf-8")
        self.logger.status(f"Coverage report written: {out}")
        return out


def generate_column_coverage_report(
    config_dir: Optional[str | Path] = None,
    output_path: Optional[str | Path] = None,
) -> Path:
    """Convenience runner for T1.311 — writes the coverage report."""
    loader = BaseAssetLoader(config_dir=config_dir)
    try:
        report_path = loader.write_coverage_report(
            output_path or (loader._repo_root / "eks" / "workplan" / "reports"
                            / "rp_p1.2_asset_column_coverage.md")
        )
    finally:
        loader.close()
    return report_path


if __name__ == "__main__":
    # Spike demo (T1.309): load the two supported sheets and emit a coverage
    # report (T1.311).
    demo = BaseAssetLoader()
    try:
        demo.logger.status(f"Workbook: {demo.workbook_path}")
        demo.logger.status(f"Sheets: {demo.sheet_names}")
        for sheet, atype in (("Equipment", "AT_EQPMP"), ("Motor", "AT_MOTOR")):
            res = demo.load_sheet(sheet, tag_type=atype, limit=5)
            demo.logger.info(
                f"{sheet} -> {atype}: {len(res.records)} records, "
                f"{len(res.unmapped_columns)} unmapped columns"
            )
            for rec in res.records[:1]:
                demo.logger.info(f"  sample: {rec['source_meta']}")
                demo.logger.info(f"  item_core keys: {sorted(rec['item_core'])}")
                demo.logger.info(f"  health: {rec['asset_health']['grade']} "
                                 f"completeness={rec['asset_health']['completeness']}")
            for warning in res.warnings[:5]:
                demo.logger.warning(warning)
        report = generate_column_coverage_report()
        demo.logger.status(f"Coverage report: {report}")
    finally:
        demo.close()
